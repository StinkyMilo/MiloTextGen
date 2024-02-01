"""
Spec - 
top row is column labels
optional weight item to the right of each
anything in brackets should be a column label

TODO:

download wordnet corpi as needed

Completed:

Variables:
[a=rule] defines a variable a to be equal to a random selection from rule.
[a] references variable a again, exactly how it was generated before.

Commands:
[rule.a] puts a or an before a noun at the beginning of a produced string (consider using en.article instead)
[rule.cap] capitalizes the first letter of produced string. All other capitalization left the same
[rule.ed] produces past tense of a verb at the end of a produced string (use en.verb.past)
[rule.s] produces plural of noun at the end of a produced string (en.verb.plural)
[rule.ing] produces gerund of verb at the end of a produced string (use en.verb.present_participle)
[rule.er] produces comparative
[rule.est] produces superlative
[rule.1p] produces 1st person version of a verb at the end of a produced string
[rule.2p] produces 2nd person version of a verb at the end of a produced string
[rule.3p] produces 3rd person version of a verb at the end of a produced string

Compatibility with:
    csv
    tsv
    xlsx

Make into a class, constructor has a filename. Format inferred from filename
    
API:
    function to load from file
    function to generate one item
    function to generate X items

"""
from random import randint
import re
import nltk
from nltk.corpus import cmudict
from nltk.corpus import words as allwords
from pattern.en import pluralize, conjugate, comparative, superlative
import openpyxl as opy
# from pattern.en.wordlist import ACADEMIC, BASIC, PROFANITY, TIME

all_words = set(allwords.words())

def weighted_choice(entry):
    val = randint(0,entry[2]-1)
    sum=0
    for i in range(0,len(entry[1])):
        sum+=entry[1][i]
        if val < sum:
            return entry[0][i]
    return None

try:
    pronunciations = cmudict.dict()
except:
    nltk.download('cmudict')
    pronunciations = cmudict.dict()

exceptions = {"a":set(),"an":set()}

def first_word(string):
    if " " not in string:
        return string
    return string[:string.index(" ")-1].lower()

def last_word(string):
    return string.split(" ")[-1]

def a_or_an(string):
    # en article doesn't work well. Trying this instead
    if len(string)==0:
        return "a"
    word = first_word(string)
    if word in exceptions['a']:
        return 'a'
    if word in exceptions['an']:
        return 'an'
    pron = pronunciations.get(word)
    if pron is None:
        return "an" if word[0] in {'a','e','i','o','u'} else 'a'
    # from https://stackoverflow.com/questions/20336524/verify-correct-use-of-a-and-an-in-english-texts-python
    for syllables in pron:
        return "an" if syllables[0][-1].isdigit() else "a"
    
def match_case(word1,word2):
    if word1.upper()==word1:
        return word2.upper()
    output = ""
    for i in range(len(word2)):
        next_char = word2[i]
        upper = i < len(word1) and word1[i]==word1[i].upper()
        if upper:
            output+=next_char.upper()
        else:
            output+=next_char.lower()
    return output

def make_plural(string):
    wordl = string.split(" ")
    word = wordl[-1].lower()
    # This library takes forever. I'll manually figure it out.
    if word in all_words:
        # TODO Fix like below.
        wordl[-1] = match_case(wordl[-1],pluralize(word))
        return " ".join(wordl)
    # Adapted from https://www.teachstarter.com/au/teaching-resource/rules-for-plurals-s-es-ies-ves/
    
    # [^aeiou](y)$ vowel + y = replace y with ies
    # (s|ch|sh|x|z)$ = add "es"
    # fe?$ = replace captured group with ves
    # otherwise add s
    # print("Pluralizing fake word",word)
    ies = re.search(r"[^aeiou](y)$",word)
    if ies is not None:
        plural = word[:len(word)-1]+"ies"
    else:
        es = re.search(r"(s|ch|sh|x|z)$",word)
        if es is not None:
            plural = word + "es"
        else:
            ves = re.search(r"fe?$",word)
            if ves is not None:
                plural = word[:ves.span(0)]+"ves"
            else:
                plural = word+"s"
    # TODO: Just replace the last word
    wordl[-1]=match_case(wordl[-1],plural)
    return " ".join(wordl)

def make_past(string):
    wordl = string.split(" ")
    word = wordl[-1].lower()
    # print("Conjugating",word)
    if word in all_words:
        try:
            p1 = conjugate(word,'1sg')
            p2 = conjugate(word,'2sg')
            p3 = conjugate(word,'3sg')
            pl = conjugate(word,'pl')
            if word == p1:
                new_word = conjugate(word,'1sgp')
            elif word==p2:
                new_word = conjugate(word,'2sgp')
            elif word==p3:
                new_word = conjugate(word,'3sgp')
            elif word==pl:
                new_word = conjugate(word,'ppl')
            wordl[-1]=match_case(wordl[-1],new_word)
            return " ".join(wordl)
        except:
            pass
    # Adapted from https://www.gingersoftware.com/content/grammar-rules/verbs/the-past-simple-tense
    if re.match(r"[^aeiou]y$",word):
        new_word = word[:len(word)-1] + "ied"
    elif re.match(r"e$",word):
        new_word = word+"d"
    elif re.match(r"[^aeiou][aeiou][^aeiouwxy]$",word):
        # Doesn't work if the stress isn't on the last syllable, but determining stress would be
        # too much work.
        new_word = word + word[-1] + "ed"
    else:
        new_word = word+"ed"
    wordl[-1] = match_case(wordl[-1],new_word)
    return " ".join(wordl)

def make_participle(string):
    wordl = string.split(" ")
    word = wordl[-1].lower()
    if word in all_words:
        try:
            part = conjugate(word,'part')
            wordl[-1] = match_case(wordl[-1],part)
            return " ".join(wordl)
        except:
            pass
    if re.match(r"[^aeiou][aeiou][^aeiouwxy]$",word):
        # Doesn't work if the stress isn't on the last syllable, but determining stress would be
        # too much work.
        new_word = word + word[-1] + "ing"
    else:
        new_word = word + "ing"
    wordl[-1] = match_case(wordl[-1],new_word)
    return " ".join(wordl)

def make_comparative(string):
    wordl = string.split(" ")
    word = wordl[-1].lower()
    if word in all_words:
        try:
            comp = comparative(word)
            wordl[-1] = match_case(wordl[-1],comp)
            return " ".join(wordl)
        except:
            pass
    if re.match(r"[^aeiou][aeiou][^aeiouwxy]$",word):
        # Doesn't work if the stress isn't on the last syllable, but determining stress would be
        # too much work.
        new_word = word + word[-1] + "er"
    else:
        new_word = word + "er"
    wordl[-1] = match_case(wordl[-1],new_word)
    return " ".join(wordl)

def make_superlative(string):
    wordl = string.split(" ")
    word = wordl[-1].lower()
    if word in all_words:
        try:
            soup = superlative(word)
            wordl[-1] = match_case(wordl[-1],soup)
            return " ".join(wordl)
        except:
            pass
    if re.match(r"[^aeiou][aeiou][^aeiouwxy]$",word):
        # Doesn't work if the stress isn't on the last syllable, but determining stress would be
        # too much work.
        new_word = word + word[-1] + "est"
    else:
        new_word = word + "est"
    wordl[-1] = match_case(wordl[-1],new_word)
    return " ".join(wordl)

def make_person(string,p):
    wordl = string.split(" ")
    word = wordl[-1].lower()
    if word in all_words:
        try:
            res = conjugate(word,p+'sg')
            wordl[-1] = match_case(wordl[-1],res)
            return " ".join(wordl)
        except:
            pass
    # No change if not a real word. No guesses for fake words on this one.
    return string

def apply_commands(commands,val):
    for command in commands:
        if command=="cap" and len(val) > 0:
            val = val[0].upper() + (val[1:] if len(val)>1 else "")
        elif command=="a":
            val = a_or_an(val) + " " + val
        elif command=="s":
            val = make_plural(val)
        elif command=="ed":
            val = make_past(val)
        elif command=="ing":
            val = make_participle(val)
        elif command=="er":
            val = make_comparative(val)
        elif command=="est":
            val = make_superlative(val)
        elif command=="1p":
            val = make_person(val,'1')
        elif command=="2p":
            val = make_person(val,'2')
        elif command=="3p":
            val = make_person(val,'3')
    return val

class Generator:
    def __init__(self,filename,worksheet_name=None):
        extension = re.search(r"\..*$",filename).group()[1:]
        if extension == "xlsx" or extension == "xls":
            items = []
            workbook = opy.load_workbook(filename)
            if worksheet_name is None:
                worksheet = workbook[workbook.sheetnames[0]]
            else:
                worksheet = workbook[worksheet_name]
            for y in range(1,worksheet.max_row+1):
                items.append([])
                for x in range(1,worksheet.max_column+1):
                    cell = worksheet.cell(x,y)
                    # For now, just convert to string then back to int. Not a big deal.
                    items[-1].append('' if cell.internal_value is None else str(cell.internal_value))
            self.items = items
        else:
            if extension=="tsv":
                delimeter = r"\t"
            elif extension=="csv":
                delimeter = r","
            items = [re.split(delimeter,i) for i in open(filename,'r').read().split("\n")]
            items_t = []
            for i in range(0,len(items)):
                for j in range(0,len(items[i])):
                    if items[i][j]=='':
                        continue
                    while len(items_t) <= j:
                        items_t.append([])
                    while len(items_t[j]) <= i:
                        items_t[j].append(None)
                    items_t[j][i]=items[i][j]
            self.items=items_t
            items = items_t
        # texts, weights, total weight
        self.grammar = {}
        i=0
        while i < len(self.items):
            column = self.items[i]
            self.grammar[column[0]]=[column[1:]]
            if i+1 >= len(self.items) or self.items[i+1][0]!='weight':
                i+=1
                self.grammar[column[0]].append([1]*(len(column)-1))
                self.grammar[column[0]].append(len(column)-1)
                continue
            weight = [int(i) for i in items[i+1][1:] if len(i)>0]
            self.grammar[column[0]].append(weight)
            self.grammar[column[0]].append(sum(weight))
            i+=2
        # print(self.grammar)

    # Not working suddenly! Replacements aren't doing what they should
    def generate(self,variables={},from_rule='root'):
        text_value = weighted_choice(self.grammar[from_rule])
        # print("Initial value:",text_value)
        match = re.search(r"\[[^\[\]]*\]",text_value)
        # print(match)
        while match:
            match_text = match.group()[1:len(match.group())-1]
            split_by_equals = match_text.split("=")
            split_by_dot = match_text.split(".")
            commands = []
            if len(split_by_dot)>1:
                match_text = split_by_dot[0]
                commands = split_by_dot[1:]
            span = match.span()
            val = ""
            if len(split_by_equals)==1:
                if match_text in variables:
                    # print("Getting value of variable",match_text)
                    val = apply_commands(commands,variables[match_text])
                else:
                    # print("Generating text for variable",match_text)
                    val = apply_commands(commands,self.generate(variables,match_text))
            else:
                # Declare variable
                # print("Declaring variable",split_by_equals,"with value determined by rule",split_by_equals[1])
                val = apply_commands(commands,self.generate(variables,split_by_equals[1].split(".")[0]))
                variables[split_by_equals[0]]=val
            
            text_value = text_value[:span[0]]+val+text_value[span[1]:]
            match = re.search(r"\[[^\[\]]*\]",text_value)
            # print(match)
            # print("New Value: " + text_value)
        text_value = text_value.replace("\\n","\n")
        return text_value
    
    def generate_multi(self,num):
        output = []
        for i in range(0,num):
            output.append(self.generate())
        return output

gen = Generator("XlsxTest.xlsx")
print(gen.generate_multi(100))